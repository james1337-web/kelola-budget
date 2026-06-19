from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Query, Header, Response
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import requests
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import mm

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Object Storage Configuration
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "budget-app"
storage_key = None
from fastapi.middleware.cors import CORSMiddleware

# Create the main app
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://kelola-budget.onrender.com"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

api_router = APIRouter(prefix="/api")

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

# Storage Helper Functions
def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    try:
        resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
        resp.raise_for_status()
        storage_key = resp.json()["storage_key"]
        logger.info("Storage initialized successfully")
        return storage_key
    except Exception as e:
        logger.error(f"Storage init failed: {e}")
        raise

def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120
    )
    resp.raise_for_status()
    return resp.json()

def get_object(path: str) -> tuple:
    key = init_storage()
    resp = requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

# Models
class DivisionBudget(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    divisi: str
    amount: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DivisionBudgetCreate(BaseModel):
    divisi: str
    amount: float

class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    divisi: str
    tanggal: str
    nominal: float
    kategori: str
    keterangan: str
    bukti_path: Optional[str] = None
    bukti_filename: Optional[str] = None
    is_deleted: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ExpenseCreate(BaseModel):
    divisi: str
    tanggal: str
    nominal: float
    kategori: str
    keterangan: str
    bukti_path: Optional[str] = None
    bukti_filename: Optional[str] = None

class ExpenseUpdate(BaseModel):
    tanggal: Optional[str] = None
    nominal: Optional[float] = None
    kategori: Optional[str] = None
    keterangan: Optional[str] = None
    divisi: Optional[str] = None
    bukti_path: Optional[str] = None
    bukti_filename: Optional[str] = None

class DivisionStats(BaseModel):
    divisi: str
    budget_awal: float
    total_pengeluaran: float
    saldo_tersisa: float
    persentase_terpakai: float

class DashboardStats(BaseModel):
    total_budget: float
    total_pengeluaran: float
    total_saldo: float
    divisions: List[DivisionStats]

class MonthlyExpense(BaseModel):
    bulan: str
    total: float

class CategoryExpense(BaseModel):
    kategori: str
    total: float

class StatsResponse(BaseModel):
    monthly: List[MonthlyExpense]
    by_category: List[CategoryExpense]

# API Endpoints
@api_router.get("/")
async def root():
    return {"message": "Budget App API - Multi Division"}

@api_router.post("/budgets", response_model=DivisionBudget)
async def create_or_update_budget(input: DivisionBudgetCreate):
    if input.amount < 0:
        raise HTTPException(status_code=400, detail="Budget tidak boleh negatif")
    
    existing = await db.division_budgets.find_one({"divisi": input.divisi}, {"_id": 0})
    
    if existing:
        update_data = {
            "amount": input.amount,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.division_budgets.update_one({"id": existing["id"]}, {"$set": update_data})
        updated = await db.division_budgets.find_one({"id": existing["id"]}, {"_id": 0})
        if isinstance(updated['created_at'], str):
            updated['created_at'] = datetime.fromisoformat(updated['created_at'])
        if isinstance(updated['updated_at'], str):
            updated['updated_at'] = datetime.fromisoformat(updated['updated_at'])
        return DivisionBudget(**updated)
    else:
        budget_obj = DivisionBudget(divisi=input.divisi, amount=input.amount)
        doc = budget_obj.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.division_budgets.insert_one(doc)
        return budget_obj

@api_router.get("/budgets/{divisi}", response_model=DivisionBudget)
async def get_division_budget(divisi: str):
    budget = await db.division_budgets.find_one({"divisi": divisi}, {"_id": 0})
    if not budget:
        return DivisionBudget(divisi=divisi, amount=0)
    
    if isinstance(budget['created_at'], str):
        budget['created_at'] = datetime.fromisoformat(budget['created_at'])
    if isinstance(budget['updated_at'], str):
        budget['updated_at'] = datetime.fromisoformat(budget['updated_at'])
    return DivisionBudget(**budget)

@api_router.get("/budgets", response_model=List[DivisionBudget])
async def get_all_budgets():
    budgets = await db.division_budgets.find({}, {"_id": 0}).to_list(100)
    for budget in budgets:
        if isinstance(budget['created_at'], str):
            budget['created_at'] = datetime.fromisoformat(budget['created_at'])
        if isinstance(budget['updated_at'], str):
            budget['updated_at'] = datetime.fromisoformat(budget['updated_at'])
    return budgets

@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    try:
        ext = file.filename.split(".")[-1] if "." in file.filename else "bin"
        path = f"{APP_NAME}/bukti/{uuid.uuid4()}.{ext}"
        data = await file.read()
        result = put_object(path, data, file.content_type or "application/octet-stream")
        
        return {
            "path": result["path"],
            "filename": file.filename,
            "size": result["size"]
        }
    except Exception as e:
        logger.error(f"Upload failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/files/{path:path}")
async def download_file(path: str):
    try:
        data, content_type = get_object(path)
        return Response(content=data, media_type=content_type)
    except Exception as e:
        logger.error(f"Download failed: {e}")
        raise HTTPException(status_code=404, detail="File not found")

@api_router.post("/expenses", response_model=Expense)
async def create_expense(input: ExpenseCreate):
    if input.nominal <= 0:
        raise HTTPException(status_code=400, detail="Nominal harus lebih dari 0")
    expense_obj = Expense(**input.model_dump())
    doc = expense_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.expenses.insert_one(doc)
    return expense_obj

@api_router.get("/expenses", response_model=List[Expense])
async def get_expenses(divisi: Optional[str] = None):
    query = {"is_deleted": False}
    if divisi:
        query["divisi"] = divisi
    expenses = await db.expenses.find(query, {"_id": 0}).to_list(1000)
    for expense in expenses:
        if isinstance(expense['created_at'], str):
            expense['created_at'] = datetime.fromisoformat(expense['created_at'])
        if isinstance(expense['updated_at'], str):
            expense['updated_at'] = datetime.fromisoformat(expense['updated_at'])
    return expenses

@api_router.put("/expenses/{expense_id}", response_model=Expense)
async def update_expense(expense_id: str, input: ExpenseUpdate):
    existing = await db.expenses.find_one({"id": expense_id, "is_deleted": False}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    if "nominal" in update_data and update_data["nominal"] <= 0:
        raise HTTPException(status_code=400, detail="Nominal harus lebih dari 0")
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.expenses.update_one({"id": expense_id}, {"$set": update_data})
    updated = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    
    if isinstance(updated['created_at'], str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    if isinstance(updated['updated_at'], str):
        updated['updated_at'] = datetime.fromisoformat(updated['updated_at'])
    return Expense(**updated)

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str):
    result = await db.expenses.update_one(
        {"id": expense_id},
        {"$set": {"is_deleted": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"message": "Expense deleted successfully"}

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats():
    budgets = await db.division_budgets.find({}, {"_id": 0}).to_list(100)
    expenses = await db.expenses.find({"is_deleted": False}, {"_id": 0}).to_list(1000)
    
    divisions_stats = []
    total_budget = 0
    total_pengeluaran = 0
    
    for divisi_name in ["REGIS", "BLAST", "SEO"]:
        budget = next((b for b in budgets if b["divisi"] == divisi_name), None)
        budget_awal = budget["amount"] if budget else 0
        
        divisi_expenses = [e for e in expenses if e["divisi"] == divisi_name]
        divisi_total = sum(exp["nominal"] for exp in divisi_expenses)
        
        saldo_tersisa = budget_awal - divisi_total
        persentase_terpakai = (divisi_total / budget_awal * 100) if budget_awal > 0 else 0
        
        divisions_stats.append(DivisionStats(
            divisi=divisi_name,
            budget_awal=budget_awal,
            total_pengeluaran=divisi_total,
            saldo_tersisa=saldo_tersisa,
            persentase_terpakai=persentase_terpakai
        ))
        
        total_budget += budget_awal
        total_pengeluaran += divisi_total
    
    return DashboardStats(
        total_budget=total_budget,
        total_pengeluaran=total_pengeluaran,
        total_saldo=total_budget - total_pengeluaran,
        divisions=divisions_stats
    )

@api_router.get("/expenses/stats/{divisi}", response_model=StatsResponse)
async def get_expense_stats(divisi: str):
    expenses = await db.expenses.find({"divisi": divisi, "is_deleted": False}, {"_id": 0}).to_list(1000)
    
    # Monthly stats
    monthly_data = {}
    for exp in expenses:
        try:
            date_parts = exp["tanggal"].split("-")
            if len(date_parts) >= 2:
                month_key = f"{date_parts[0]}-{date_parts[1]}"
                monthly_data[month_key] = monthly_data.get(month_key, 0) + exp["nominal"]
        except (ValueError, KeyError, AttributeError):
            pass
    
    monthly = [MonthlyExpense(bulan=k, total=v) for k, v in sorted(monthly_data.items())]
    
    # Category stats
    category_data = {}
    for exp in expenses:
        cat = exp["kategori"]
        category_data[cat] = category_data.get(cat, 0) + exp["nominal"]
    
    by_category = [CategoryExpense(kategori=k, total=v) for k, v in category_data.items()]
    
    return StatsResponse(monthly=monthly, by_category=by_category)

@api_router.get("/export/excel/{divisi}")
async def export_excel(divisi: str):
    expenses = await db.expenses.find({"divisi": divisi, "is_deleted": False}, {"_id": 0}).to_list(1000)
    budget = await db.division_budgets.find_one({"divisi": divisi}, {"_id": 0})
    budget_awal = budget["amount"] if budget else 0
    total_pengeluaran = sum(exp["nominal"] for exp in expenses)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{divisi}"
    
    # Header style
    header_fill = PatternFill(start_color="183623", end_color="183623", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"LAPORAN BUDGET - {divisi}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')
    
    ws['A2'] = f"Budget Awal: Rp {budget_awal:,.0f}".replace(",", ".")
    ws['A3'] = f"Total Pengeluaran: Rp {total_pengeluaran:,.0f}".replace(",", ".")
    ws['A4'] = f"Saldo Tersisa: Rp {budget_awal - total_pengeluaran:,.0f}".replace(",", ".")
    
    # Headers
    headers = ["Tanggal", "Nominal", "Kategori", "Keterangan", "Bukti"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # Data
    for idx, exp in enumerate(expenses, 7):
        ws.cell(row=idx, column=1, value=exp["tanggal"]).border = thin_border
        ws.cell(row=idx, column=2, value=f"Rp {exp['nominal']:,.0f}".replace(",", ".")).border = thin_border
        ws.cell(row=idx, column=3, value=exp["kategori"]).border = thin_border
        ws.cell(row=idx, column=4, value=exp["keterangan"]).border = thin_border
        ws.cell(row=idx, column=5, value=exp["bukti_filename"] or "-").border = thin_border
    
    # Total
    total_row = len(expenses) + 7
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=f"Rp {total_pengeluaran:,.0f}".replace(",", ".")).font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 30
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=laporan_{divisi}.xlsx"}
    )

@api_router.get("/export/pdf/{divisi}")
async def export_pdf(divisi: str):
    expenses = await db.expenses.find({"divisi": divisi, "is_deleted": False}, {"_id": 0}).to_list(1000)
    budget = await db.division_budgets.find_one({"divisi": divisi}, {"_id": 0})
    budget_awal = budget["amount"] if budget else 0
    total_pengeluaran = sum(exp["nominal"] for exp in expenses)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#183623'),
        spaceAfter=30,
        alignment=1
    )
    elements.append(Paragraph(f"LAPORAN BUDGET - {divisi}", title_style))
    elements.append(Spacer(1, 12))
    
    # Budget info
    elements.append(Paragraph(f"Budget Awal: Rp {budget_awal:,.0f}".replace(",", "."), styles['Normal']))
    elements.append(Paragraph(f"Total Pengeluaran: Rp {total_pengeluaran:,.0f}".replace(",", "."), styles['Normal']))
    elements.append(Paragraph(f"Saldo Tersisa: Rp {budget_awal - total_pengeluaran:,.0f}".replace(",", "."), styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table
    data = [['Tanggal', 'Nominal', 'Kategori', 'Keterangan']]
    for exp in expenses:
        data.append([
            exp["tanggal"],
            f"Rp {exp['nominal']:,.0f}".replace(",", "."),
            exp["kategori"],
            exp["keterangan"][:30] + "..." if len(exp["keterangan"]) > 30 else exp["keterangan"]
        ])
    
    table = Table(data, colWidths=[60, 80, 80, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#183623')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=laporan_{divisi}.pdf"}
    )

@app.on_event("startup")
async def startup():
    try:
        init_storage()
        logger.info("Application started successfully")
    except Exception as e:
        logger.error(f"Startup failed: {e}")

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
